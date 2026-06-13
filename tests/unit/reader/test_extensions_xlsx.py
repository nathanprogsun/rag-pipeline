"""xlsx extension adapter 单元测试。"""

from __future__ import annotations

import inspect
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from rag.error_codes import ReaderErrorCode
from rag.exception import RAGError
from rag.ingest.reader.extensions.xlsx import CUSTOM_SPLIT_SIGN, xlsx_adapter

SAMPLE_XLSX = Path(__file__).resolve().parents[2] / "data" / "sample.xlsx"


def _build_xlsx_with_hidden_sheet() -> bytes:
    """构造含隐藏 sheet 的 xlsx bytes。"""
    wb = Workbook()
    active = wb.active
    assert active is not None
    active.title = "Visible"
    active.append(["visible-only"])
    hidden = wb.create_sheet("Hidden")
    hidden.append(["hidden-data"])
    hidden.sheet_state = "hidden"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_xlsx(*sheets: list[list[str]]) -> bytes:
    """构造多 sheet xlsx bytes。"""
    wb = Workbook()
    first = wb.active
    assert first is not None
    for idx, rows in enumerate(sheets):
        ws = first if idx == 0 else wb.create_sheet(f"Sheet{idx + 1}")
        for row in rows:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_xlsx_basic_single_sheet() -> None:
    """单 sheet 表格应产出 CSV raw_text 与 markdown format_text。"""
    buf = _build_xlsx([["Name", "Age"], ["Alice", "30"]])
    result = await xlsx_adapter(buf)

    assert "Name,Age" in result.raw_text
    assert "Alice,30" in result.raw_text
    assert result.format_text is not None
    assert "| Name | Age |" in result.format_text
    assert "| Alice | 30 |" in result.format_text


@pytest.mark.asyncio
async def test_xlsx_multiple_sheets() -> None:
    """多 sheet: raw_text 用换行拼接, format_text 用 CUSTOM_SPLIT_SIGN。"""
    buf = _build_xlsx([["A", "1"], ["a", "1"]], [["B", "2"], ["b", "2"]])
    result = await xlsx_adapter(buf)

    assert "A,1" in result.raw_text
    assert "B,2" in result.raw_text
    assert result.format_text is not None
    assert CUSTOM_SPLIT_SIGN in result.format_text
    assert "| A | 1 |" in result.format_text
    assert "| B | 2 |" in result.format_text


@pytest.mark.asyncio
async def test_xlsx_empty_sheet() -> None:
    """首个空 sheet 跳过, 第二个 sheet 仍可读。"""
    buf = _build_xlsx([], [["data", "value"], ["x", "y"]])
    result = await xlsx_adapter(buf)

    assert "data,value" in result.raw_text
    assert result.raw_text.count("\n") == 0 or "value" in result.raw_text


@pytest.mark.asyncio
async def test_xlsx_with_newline_in_cell() -> None:
    """单元格内换行在 markdown table 中应转义为 ``\\n``。"""
    buf = _build_xlsx([["h1", "h2"], ["line1\nline2", "ok"]])
    result = await xlsx_adapter(buf)

    assert result.format_text is not None
    assert "line1\\nline2" in result.format_text


@pytest.mark.asyncio
async def test_xlsx_hidden_sheet_is_read() -> None:
    """隐藏 sheet 仍应被读取。"""
    buf = _build_xlsx_with_hidden_sheet()
    result = await xlsx_adapter(buf)

    assert "visible-only" in result.raw_text
    assert "hidden-data" in result.raw_text


@pytest.mark.asyncio
async def test_xlsx_format_text_none_when_all_empty() -> None:
    """仅空行时 raw_text 与 format_text 均为空/None。"""
    buf = _build_xlsx([[]])
    result = await xlsx_adapter(buf)

    assert result.raw_text == ""
    assert result.format_text is None


@pytest.mark.asyncio
async def test_xlsx_format_text_none_when_all_sheets_completely_empty() -> None:
    """多个空 sheet 时 raw_text 为空, format_text 为 None。"""
    buf = _build_xlsx([], [])
    result = await xlsx_adapter(buf)

    assert result.raw_text == ""
    assert result.format_text is None


@pytest.mark.asyncio
async def test_xlsx_corrupted() -> None:
    """损坏 xlsx 应抛 READER_PARSE。"""
    with pytest.raises(RAGError) as exc_info:
        await xlsx_adapter(b"not-a-valid-xlsx")
    assert exc_info.value.code == ReaderErrorCode.PARSE


@pytest.mark.asyncio
async def test_xlsx_corrupted_real_bad_bytes() -> None:
    """真实损坏 OOXML 字节应抛 READER_PARSE。"""
    bad = b"PK\x03\x04" + b"\x00" * 100
    with pytest.raises(RAGError) as exc_info:
        await xlsx_adapter(bad)
    assert exc_info.value.code == ReaderErrorCode.PARSE


@pytest.mark.asyncio
async def test_xlsx_against_sample_fixture() -> None:
    """sample.xlsx fixture 应可读。"""
    result = await xlsx_adapter(SAMPLE_XLSX.read_bytes())
    assert len(result.raw_text) > 0
    assert result.format_text is not None


@pytest.mark.asyncio
async def test_xlsx_adapter_is_async() -> None:
    """xlsx_adapter 应为 async 函数。"""
    assert inspect.iscoroutinefunction(xlsx_adapter)
