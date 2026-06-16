"""xlsx 格式适配器: openpyxl 抽 sheet, 同时输出 CSV 与 markdown 表格两种视图。"""

from __future__ import annotations

import logging
import zipfile
from io import BytesIO
from typing import Final

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from rag.ingest.reader.extensions.base import wrap_parse_error
from rag.ingest.reader.types import FormatReaderResult
from rag.ingest.types import DocMeta

logger = logging.getLogger(__name__)

# xlsx 标准 mime (OOXML 官方命名)。
XLSX_MIME: Final[str] = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# 多 sheet 的 markdown 表格拼接分隔符。
# 注意: 字面值与 ``chunker.rules.CUSTOM_SPLIT_SIGN`` 相同, 但语义不同 —
# chunker 那个用于 chunk 间占位, xlsx 这里用于 sheet 间分隔, 不要混用。
CUSTOM_SPLIT_SIGN: Final[str] = "-----CUSTOM-SPLIT-SIGN-----"


def _sheet_to_rows(ws: object) -> list[list[str]]:
    """把 sheet 转成 ``[[cell, ...], ...]``, ``None`` 转为空串并过滤完全空行。

    Args:
        ws: openpyxl 的 worksheet 对象。

    Returns:
        单元格二维列表。
    """
    raw_data: list[tuple[object, ...]] = list(ws.iter_rows(values_only=True))  # type: ignore[attr-defined]
    rows: list[list[str]] = []
    for row in raw_data:
        cells = ["" if v is None else str(v) for v in row]
        # 过滤"完全空行": 全部 cell 都是空串
        if all(c == "" for c in cells):
            continue
        rows.append(cells)
    return rows


def _sheet_to_csv(rows: list[list[str]]) -> str:
    """把 ``[[cell, ...], ...]`` 拼成 CSV 文本, 行内 ``,``, 行间换行。"""
    return "\n".join(",".join(row) for row in rows)


def _sheet_to_md_table(rows: list[list[str]]) -> str | None:
    """把 ``[[cell, ...], ...]`` 拼成 markdown 表格, 行数不足 2 时返回 ``None``。

    Args:
        rows: 单元格二维列表。

    Returns:
        markdown 表格字符串, 或 ``None``。
    """
    if len(rows) < 2:
        return None

    header = rows[0]
    body = rows[1:]
    width = len(header)

    lines: list[str] = [
        "| " + " | ".join(header) + " |",
        "|" + "|".join(["---"] * width) + "|",
    ]
    for row in body:
        # 列数补齐 / 截断到 header 宽度
        if len(row) < width:
            row = list(row) + [""] * (width - len(row))
        elif len(row) > width:
            row = list(row)[:width]
        escaped = [c.replace("\n", "\\n") for c in row]
        lines.append("| " + " | ".join(escaped) + " |")
    return "\n".join(lines)


async def xlsx_adapter(
    buffer: bytes,
    *,
    encoding: str = "utf-8",  # xlsx 不需要 encoding, 保留对齐签名
) -> FormatReaderResult:
    """将 xlsx 字节内容解析为 ``FormatReaderResult``, 同时提供 CSV 与 markdown 两种视图。

    Args:
        buffer: xlsx 二进制内容。
        encoding: 占位参数 (xlsx 是二进制 OOXML, 不需要文本编码),
            保留以对齐 ``FormatAdapter`` 签名。

    Returns:
        ``FormatReaderResult``:
        - ``raw_text``: 多 sheet 的 CSV 拼接 (sheet 间换行, 行内 ``,``, 无 sheet 标题)。
        - ``format_text``: 多 sheet 的 markdown 表格拼接, sheet 间用
          ``CUSTOM_SPLIT_SIGN``; 若所有 sheet 行数不足 2, 则为 ``None``。
        - ``meta.mime`` 为 xlsx 标准 mime。

    Raises:
        RAGError: ``code=READER_PARSE`` —— openpyxl 解析失败 (坏 zip / 非法 xlsx)。
    """
    del encoding  # unused; xlsx 是二进制, 不需 decode

    try:
        # data_only=True 取计算后的值, read_only=True 省内存
        wb = load_workbook(BytesIO(buffer), data_only=True, read_only=True)
    except (InvalidFileException, zipfile.BadZipFile) as e:
        raise wrap_parse_error("<buffer:xlsx>", e, "openpyxl") from e
    except Exception as e:
        # openpyxl 内部也可能抛 OSError / ValueError / KeyError 等, 统一包装
        raise wrap_parse_error("<buffer:xlsx>", e, "openpyxl") from e

    csv_chunks: list[str] = []
    md_chunks: list[str] = []

    try:
        for ws in wb.worksheets:
            rows = _sheet_to_rows(ws)

            # 所有非空 sheet 都贡献 CSV 块
            if rows:
                csv_chunks.append(_sheet_to_csv(rows))

            # 只有 len(rows) >= 2 的 sheet 才贡献 markdown 表格
            md = _sheet_to_md_table(rows)
            if md is not None:
                md_chunks.append(md)
    finally:
        # read_only 模式: 显式 close 释放内存 + 文件句柄
        wb.close()

    # 多 sheet 直接换行拼接, 无 sheet 标题
    raw_text = "\n".join(csv_chunks)
    # 多 sheet 用 CUSTOM_SPLIT_SIGN 拼接; 若全部被过滤则为 None
    format_text: str | None = CUSTOM_SPLIT_SIGN.join(md_chunks) if md_chunks else None

    return FormatReaderResult(
        raw_text=raw_text,
        format_text=format_text,
        meta=DocMeta(mime=XLSX_MIME),
        images=[],
        extras={},
    )
